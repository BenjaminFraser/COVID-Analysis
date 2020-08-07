#!/usr/bin/env python
# -*- coding: utf-8 -*-

# import dependencies and external libs
import datetime
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pmdarima as pm
import os
import xlsxwriter
import warnings

import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pandas import ExcelWriter
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.metrics import r2_score, mean_squared_error
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.svm import LinearSVR
from sklearn.ensemble import BaggingRegressor
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from statsmodels.tsa.arima_model import ARIMA
from tqdm import tqdm

# import cubic spline model classes
from cubic_splines.natural_cubic_spline import AbstractSpline, NaturalCubicSpline

sns.set_style('darkgrid')

# obtain current date and set save file name for final exported Excel File
CURRENT_DATE = datetime.datetime.today().strftime('%Y%m%d')
FILE_SAVE_NAME = f'{CURRENT_DATE}-Positive-Cases-Analysis.xlsx'

# set directory that preprocessed data is stored within
DATA_SOURCE_DIR = os.path.join(os.getcwd(), 'Preprocessed_data')

# filename containing a map of all uk local area codes
AREA_MAPPINGS_FILENAME = 'UK-code-areas.csv'

LOCATIONS = ['Wales', 'Scotland', 'England', 'Ireland', 'Northern_Ireland', 'Other']
POLYNOMIAL_FOLDER = 'polynomial_plots'
CUBIC_SPLINE_FOLDER = 'cubic_spline_plots'
ARIMA_FOLDER = 'arima_plots'

# set directories for each location
WALES_DIR = os.path.join(os.getcwd(), 'Wales')
ENGLAND_DIR = os.path.join(os.getcwd(), 'England')
SCOTLAND_DIR = os.path.join(os.getcwd(), 'Scotland')
IRELAND_DIR = os.path.join(os.getcwd(), 'Ireland')
NI_DIR = os.path.join(os.getcwd(), 'Northern_Ireland')
OTHER_DIR = os.path.join(os.getcwd(), 'Other')

# declare start dates for each dataset
WALES_START_DATE = '2020-03-11'
ENGLAND_START_DATE = '2020-03-11'
SCOTLAND_START_DATE = '2020-03-11'
IRELAND_START_DATE = '2020-03-11'
NI_START_DATE = '2020-03-20'
OTHER_START_DATE = '2020-03-20'

# set y-ax label for all plots
Y_AXIS_LABEL = "Cases per 100,000"

# settings for final excel tables formed
FIRST_COL_WIDTHS = 30
FORECAST_WIDTHS = 15
TIME_CELL_WIDTHS = 10
GRADIENT_CELL_WIDTH = 18

# form desired final order dict for our excel sheets
FINAL_SHEET_ORDER = {'England Cases Per Pop' : 0, 'Scotland Cases Per Pop' : 1, 'Wales Cases Per Pop' : 2, 
                     'Ireland Cases Per Pop' : 3, 'NI Cases Per Pop' : 4, 'Other Cases Per Pop' : 5,
                     'England CubicSpline Plots' : 6, 'England CubicSpline Model' : 7, 
                     'Scotland CubicSpline Plots' : 8, 'Scotland CubicSpline Model' : 9, 
                     'Wales CubicSpline Plots' : 10, 'Wales CubicSpline Model' : 11,
                     'Ireland CubicSpline Plots' : 12, 'Ireland CubicSpline Model' : 13, 
                     'NI CubicSpline Plots' : 14, 'NI CubicSpline Model' : 15, 
                     'Other CubicSpline Plots' : 16, 'Other CubicSpline Model' : 17,
                     'England ARIMA Plots' : 18, 'England ARIMA Model' : 19, 
                     'Scotland ARIMA Plots' : 20, 'Scotland ARIMA Model' : 21, 
                     'Wales ARIMA Plots' : 22, 'Wales ARIMA Model' : 23,
                     'Ireland ARIMA Plots' : 24, 'Ireland ARIMA Model' : 25, 
                     'NI ARIMA Plots' : 26, 'NI ARIMA Model' : 27, 
                     'Other ARIMA Plots' : 28, 'Other ARIMA Model' : 29}


def create_dirs(location):
    """ Create required dirs within each location folder """ 
    base_dir = os.path.join(os.getcwd(), location)
    polynomial_dir = os.path.join(base_dir, POLYNOMIAL_FOLDER)
    cubic_spline_dir = os.path.join(base_dir, CUBIC_SPLINE_FOLDER)
    arima_dir = os.path.join(base_dir, ARIMA_FOLDER)

    for dir_path in [polynomial_dir, cubic_spline_dir, arima_dir]:
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)


def poly_regression(X, y, location_name, poly_transformer, predict_ahead=0,
                    n_estimators=10, figsize=(7,5), save=False, save_dir=POLYNOMIAL_FOLDER, 
                    show_plot=False, yaxis_label=Y_AXIS_LABEL):
    """ Form a polynomial regression model with bootstrapped prediction intervals for the
        given input data and output labels. The function can plot to the console, and save
        to directory using the given named arguments. Returns a tuple of model results.

    Parameters
    ----------
    X : Numpy array
        Numpy array with cols as features and rows as time (daily) samples.
    y : Numpy array
        Numpy array containing the output labels for each of the inputs in X
    location_name : String
        String containing the location name (associated column name in location dataframe)
    predict_ahead : int
        Number of days to forecast ahead with the model, relative to the last day in X
    n_estimators : int
        Number of estimators to use for the bagging model for creating the prediction interval.

    Returns
    -------
    Python tuple
        A tuple containing model results: (loc_name, equation_terms, r^2, predictions, gradient)
    """
    
    model = BaggingRegressor(LinearRegression(), 
                         n_estimators=n_estimators,
                         bootstrap=True)
    
    # set x-axis vals, obtain polynomial feats, and fit to model
    x_range = np.array(range(1, len(X) + 1 + predict_ahead)).reshape(-1, 1)
    X_poly = poly_transformer.transform(X)
    model.fit(X_poly, y)
    
    plt.figure(figsize=figsize)

    # make predictions using each individual base regressor
    for m in model.estimators_:
        predictions = m.predict(poly_transformer.transform(x_range))
        predictions = np.maximum(0.0, predictions)                      
        plt.plot(x_range, predictions, color='lightgrey', 
                 alpha=0.5, zorder=1)
    
    # find equation of base regression line
    base_model = LinearRegression().fit(X_poly, y)
    model_equation = np.append(base_model.coef_[1:], base_model.intercept_)
    equation_terms = [f"{coefficient:.4f}" for coefficient in model_equation]
    
    # plot original data
    plt.scatter(X[:, 0], y, marker='o', color='tab:blue', 
                zorder=4, label='Original Data')

    # Bagging model predictions and plot
    bagging_preds = model.predict(poly_transformer.transform(x_range))
    bagging_preds = np.maximum(0.0, bagging_preds)
    plt.plot(x_range, bagging_preds, color='tab:red', 
             zorder=5, label='Polynomial (n=3)')
    
    # add predicted values to global polyreg results
    polyreg_preds_dict[location_name] = (x_range, bagging_preds)
    
    # calculate r^2  score for our model
    r2_result = r2_score(y, bagging_preds[:-predict_ahead])
    
    # calculate final gradient (during forecasts)
    y1, y2 = bagging_preds[-2], bagging_preds[-1]
    x1, x2 = x_range[-2], x_range[-1]
    end_gradient = (y2 - y1) / (x2 - x1)
    
    # save model name, equation co-efficients and predictions
    model_results = (location_name, *equation_terms, f"{r2_result:.4f}", 
                     *bagging_preds[-predict_ahead:], *end_gradient)
    
    plt.title(f"Cases in {location_name}", weight='bold', size=14)
    plt.xlabel("Days", weight='bold', size=12)
    plt.ylabel(yaxis_label, weight='bold', size=12)
    
    # set y-axis to appropriate values
    if np.max(y) < 5:
        plt.ylim(-3.0, 5.0)
    else:
        plt.ylim(-1.0)

    # save plot if selected
    plt.legend(loc='best')
    if save:
        plt.savefig(f"{save_dir}/{location_name}.png", format='png')

    # display plot if selected
    if show_plot:
        plt.show()
    else:
        plt.close()
    
    return model_results


def location_polynomials(location_df, polynomial_order=3, predict_ahead=5, 
                         save=True, save_dir=POLYNOMIAL_FOLDER, show_plot=False,
                         yaxis_label=Y_AXIS_LABEL):
    """ Plot polynomial reg models for all location cols in given dataframe 

    Parameters
    ----------
    location_df : Pandas DataFrame
        DataFrame containing time-series indexed rows of COVID-19 cases for each location column.
    polynomial_order : int
        Integer representing the polynomial order to use for the model, default cubic (3).
    location_name : String
        String containing the location name (associated column name in location dataframe)
    predict_ahead : int
        Number of days to forecast ahead with the model, relative to last date in location_df.

    Returns
    -------
    Pandas DataFrame
        A DataFrame containing all returned model results for each region/area in location df.
    """
    
    all_results = []
    
    # create polynomial transformer
    X = location_df.index.values.reshape(-1, 1)
    poly = PolynomialFeatures(degree=polynomial_order)
    poly.fit_transform(X)
    
    # apply polynomial reg to all locations (columns)
    for column in location_df.columns:
        
        # save results / details of each regression
        results = poly_regression(X, location_df[column].values, 
                                  location_name=str(column), poly_transformer=poly, 
                                  predict_ahead=predict_ahead, save=save, save_dir=save_dir,
                                  show_plot=show_plot, yaxis_label=yaxis_label)
        
        all_results.append(results)
    
    # strings for prediction columns of dataframe
    days = [f"+{i} Day Forecast" for i in range(1, predict_ahead + 1)]

    # form dataframe of all model results
    polyreg_df = pd.DataFrame(all_results, 
                              columns=['Location', 'x^3', 
                                       'x^2', 'x', 'Intercept', 
                                       'R2 Score', *days, 
                                       'Forecast Gradient'])
    
    polyreg_df.set_index('Location', inplace=True, drop=True)
    
    return polyreg_df



def get_natural_cubic_spline_model(x, y, minval=None, maxval=None, n_knots=None, knots=None):
    """ Get a natural cubic spline model for the data.

    For the knots, give (a) `knots` (as an array) or (b) minval, maxval and n_knots.

    If the knots are not directly specified, the resulting knots are equally
    space within the *interior* of (max, min).  That is, the endpoints are
    *not* included as knots.

    Parameters
    ----------
    x: np.array of float
        The input data
    y: np.array of float
        The outpur data
    minval: float 
        Minimum of interval containing the knots.
    maxval: float 
        Maximum of the interval containing the knots.
    n_knots: positive integer 
        The number of knots to create.
    knots: array or list of floats 
        The knots.

    Returns
    --------
    model: a model object
        The returned model will have following method:
        - predict(x):
            x is a numpy array. This will return the predicted y-values.
    """

    if knots:
        spline = NaturalCubicSpline(knots=knots)
    else:
        spline = NaturalCubicSpline(max=maxval, min=minval, n_knots=n_knots)

    p = Pipeline([
        ('nat_cubic', spline),
        ('regression', LinearRegression(fit_intercept=True))
    ])

    p.fit(x, y)

    return p



def nat_cubic_spline_regr(dataframe, column_name, n_knots, predict_ahead=0, n_estimators=250, 
                          figsize=(7,5), save=False, save_dir=CUBIC_SPLINE_FOLDER, 
                          show_plot=False, yaxis_label=Y_AXIS_LABEL, plot_poly=False):
    """ Perform natural cubic spline regression on given dataframe col """
    
    # set current input features and output labels
    x = dataframe.index.values.reshape(-1, 1)
    y = dataframe[column_name].values.ravel()
    
    # obtain natural cubic spline model to feed our bagging ensemble
    regr_model = get_natural_cubic_spline_model(x, y, minval=min(x), 
                                                maxval=max(x), n_knots=n_knots)
    
    # form n individual regressors based on regr_model
    model = BaggingRegressor(regr_model, n_estimators=n_estimators, bootstrap=True)
    
    # set total range of x-axis values (including trg + forecasts)
    x_range = np.array(range(1, len(x) + 1 + predict_ahead)).reshape(-1, 1)
    
    # fit data to our bagging model
    model.fit(x, y)
    
    # set plot fig for visualising bagging preds and pred interval
    plt.figure(figsize=figsize)
    
    # make predictions using each individual base regressor
    for m in model.estimators_:
        predictions = m.predict(x_range)
        plt.plot(x_range, predictions, color='lightgrey', alpha=0.5, zorder=1)
    
    # if scotland - modify to plot original uncorrected data
    if column_name in SCOTLAND_AREAS:
        # set current input features and output labels
        x_original = scotland_daily_df_original.index.values.reshape(-1, 1)
        y_original = scotland_daily_df_original[column_name].values.ravel()
        plt.scatter(x_original, y_original, marker='o', color='tab:blue', 
                    zorder=4, label='Original data')
    
    # if Northern Ireland - modify to plot original uncorrected data
    elif column_name in NI_AREAS:
        # set current input features and output labels
        x_original = ni_daily_df_original.index.values.reshape(-1, 1)
        y_original = ni_daily_df_original[column_name].values.ravel()
        plt.scatter(x_original, y_original, marker='o', color='tab:blue', 
                    zorder=4, label='Original data')
    
    # otherwise - plot original data as normal
    else:
        plt.scatter(x, y, marker='o', color='tab:blue', zorder=4, label='Original data')

    # Bagging model predictions
    bagging_preds = model.predict(x_range)
    bagging_preds = np.maximum(0.0, bagging_preds)
    plt.plot(x_range, bagging_preds, color='tab:red', 
             zorder=5, label=f'Natural Cubic Spline ({n_knots} knots)')
    
    
    # if selected, plot polynomial curve too
    if plot_poly:
        plt.plot(x_range, polyreg_preds_dict[column_name][1], alpha=0.5, 
                 color='tab:green', zorder=5, label='Polynomial (n=3)')

    # if plot poly not chosen - plot weighted average
    else:
        
        moving_avg = dataframe[column_name].rolling(window=10, center=True).mean()
        plt.plot(x, moving_avg, alpha=0.6, color='tab:blue', 
                 zorder=5, label='10 Day Moving Average', linestyle='--')
    
    # find R^2 score for measuring out fit on the data
    r2_result = r2_score(y, bagging_preds[:-predict_ahead])
    
    # calculate gradient on last data points
    y1, y2 = bagging_preds[-2], bagging_preds[-1]
    x1, x2 = x_range[-2], x_range[-1]
    end_gradient = (y2 - y1) / (x2 - x1)
    
    # append model name, equation co-efficients and predictions to results
    spline_results = (column_name,  f"{r2_result:.4f}", 
                      *bagging_preds[-predict_ahead:], *end_gradient)

    #plt.title(f"Cases in {column_name}\n $ R^{2} = {r2_result} $", weight='bold', size=14)
    plt.title(f"{column_name}", weight='bold', size=14)
    
    plt.xlabel("Days", weight='bold', size=12)
    plt.ylabel(yaxis_label, weight='bold', size=12)
    
    # set y-axis to appropriate values
    if np.max(y) < 5:
        plt.ylim(-1.0, 5.0)
    else:
        plt.ylim(-1.0)
        
    plt.legend(loc='best')
    if save:
        plt.savefig(f"{save_dir}/{column_name}.png", format='png')
        
    if show_plot:
        plt.show()
    else:
        plt.close()
        
    return spline_results



def ncs_regressions(location_df, n_knots=6, predict_ahead=5, save=True, 
                    save_dir=CUBIC_SPLINE_FOLDER, show_plot=False,
                    yaxis_label=Y_AXIS_LABEL, plot_poly=False):
    """ Plot natural cubic spline regressors for all location cols in given dataframe """
    
    all_results = []

    print("  Natural Cubic Spline Models:")

    # apply polynomial reg to all locations (columns)
    for column in tqdm(location_df.columns):
        
        # save results / details of each regression
        results = nat_cubic_spline_regr(location_df, column_name=column, n_knots=n_knots,  
                                        predict_ahead=predict_ahead, save=save, save_dir=save_dir,
                                        show_plot=show_plot, yaxis_label=Y_AXIS_LABEL, plot_poly=plot_poly)
        all_results.append(results)
    
    # strings for prediction columns of dataframe
    days = [f"+{i} Day Forecast" for i in range(1, predict_ahead + 1)]

    # form dataframe of all model results
    cubic_spline_df = pd.DataFrame(all_results, 
                                   columns=['Location', 'R2 Score', *days, 'Forecast Gradient'])
    
    cubic_spline_df.set_index('Location', inplace=True, drop=True)
    
    return cubic_spline_df


def plot_auto_arima(dataframe, column_name, predict_ahead=5, figsize=(7,5), 
                    save=False, training=False, show_plot=False, save_dir=ARIMA_FOLDER, 
                    yaxis_label=Y_AXIS_LABEL):
    """ Automatically find the best ARIMA model and use it to 
        forecast on our selected data """
    
    if training:
        train_df, test_df = train_test_split(dataframe, test_size=0.2)
    else:
        train_df = dataframe.copy()
    
    arima_model = pm.auto_arima(train_df[column_name].values, start_p=1, start_q=1,
                                test='adf',       # adftest for finding optimal 'd'
                                max_p=3, max_q=3, # maximum p and q
                                m=1,              # frequency of series
                                d=None,           # let model determine 'd'
                                seasonal=False,   # No Seasonality assumptions
                                start_P=0, 
                                D=0, 
                                trace=False,
                                error_action='ignore',  
                                suppress_warnings=True, 
                                stepwise=True)
    
    # Forecast using our best ARIMA model
    if training:
        n_periods = test_df.shape[0]
    else:
        n_periods = 5
        
    fc, confint = arima_model.predict(n_periods=n_periods, return_conf_int=True)
    index_of_fc = np.arange(len(train_df[column_name]), 
                            len(train_df[column_name]) + n_periods)

    # format results as series objects for plotting
    fc_series = pd.Series(fc, index=index_of_fc)
    lower_series = pd.Series(confint[:, 0], index=index_of_fc)
    upper_series = pd.Series(confint[:, 1], index=index_of_fc)

    # weight averages of the data to fill before the forecast
    #train_exp_wm = train_df.ewm(span=20, adjust=False).mean()
    #train_exp_wm = train_df.rolling(10).mean()
    train_exp_wm = train_df.rolling(window=10, win_type='gaussian').mean(std=10)
    
    # form a plot line consisting of rolling expential weighted mean + predictions
    avg_and_forecasts = np.concatenate([train_exp_wm[column_name].values, fc_series])
    
    # rectify all values less than zero (since this isnt possible)
    lower_series = np.maximum(0.0, lower_series)
    avg_and_forecasts = np.maximum(0.0, avg_and_forecasts)
    
    # calculate gradient on last data points
    y1, y2 = avg_and_forecasts[-2], avg_and_forecasts[-1]
    x1, x2 = train_df.index.values[-2], train_df.index.values[-1]
    end_gradient = (y2 - y1) / (x2 - x1)
    
    # append model name, equation co-efficients and predictions to results
    arima_results = (column_name, *avg_and_forecasts[-predict_ahead:], end_gradient)
    
    # plot results
    plt.figure(figsize=figsize)
    plt.scatter(x=train_df.index, y=train_df[column_name], 
                color='tab:blue', label='Original Data')
    plt.plot(avg_and_forecasts, color='tab:red', 
             label='Weighted Avg + ARIMA forecast')
        
    plt.fill_between(lower_series.index, 
                     lower_series, 
                     upper_series, 
                     color='k', alpha=.15)

    plt.title(f"{column_name} ARIMA Forecast", weight='bold', size=14)
    plt.xlabel("Days", weight='bold', size=12)
    plt.ylabel(Y_AXIS_LABEL, weight='bold', size=12)
    
    # set y-axis to appropriate values
    if np.max(train_df[column_name].values) < 5:
        plt.ylim(-1.0, 5.0)
    else:
        plt.ylim(-1.0)
    
    plt.legend(loc='best')
    if save:
        plt.savefig(f"{save_dir}/{column_name}_ARIMA.png", format='png')
    if show_plot:
        plt.show()
    else:
        plt.close()
    
    return arima_results
    
    
def arima_forecasts(location_df, predict_ahead=5, training=False, save=True, 
                    save_dir=ARIMA_FOLDER, show_plot=False, yaxis_label=Y_AXIS_LABEL):
    """ Plot arima forecasts for all location cols in given dataframe """
    
    all_results = []
    
    print("  ARIMA Models:")

    # apply arima to all locations (columns)
    for column in tqdm(location_df.columns):
        
        # save results / details of each regression
        results = plot_auto_arima(location_df, column_name=column,
                                  predict_ahead=predict_ahead, save=save, 
                                  save_dir=save_dir, show_plot=show_plot, 
                                  yaxis_label=Y_AXIS_LABEL, training=training)
        all_results.append(results)
    
    # strings for prediction columns of dataframe
    days = [f"+{i} Day Forecast" for i in range(1, predict_ahead + 1)]

    # form dataframe of all model results
    arima_df = pd.DataFrame(all_results, 
                                   columns=['Location', *days, 'Forecast Gradient'])
    
    arima_df.set_index('Location', inplace=True, drop=True)
    
    return arima_df


def get_model_results(loc_df, location_dir, y_axis_label):
    """ Call each model on all locations within the given dataframe. Returns three
        dataframes with the results from each model type """

    base_dir = location_dir
    polynomial_dir = os.path.join(base_dir, POLYNOMIAL_FOLDER)
    cubic_spline_dir = os.path.join(base_dir, CUBIC_SPLINE_FOLDER)
    arima_dir = os.path.join(base_dir, ARIMA_FOLDER)

    daily_df = loc_df.copy()

    # Convert date time index into a simple index of raw days
    daily_df['Day'] = range(1, daily_df.shape[0] + 1)
    daily_df.set_index('Day', drop=True, inplace=True)

    print(f"\nCreating plots for {os.path.basename(location_dir)}...")

    # obtain cubic polynomial regression (n=3) results
    polyreg_df = location_polynomials(daily_df, save=True, save_dir=polynomial_dir, 
                                      yaxis_label=y_axis_label)

    # ignore future warning messages obtain natural cubic splines
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore")
        # get cubic splines, with polynomial model (n=3) also included
        splines_df = ncs_regressions(daily_df, save=True, save_dir=cubic_spline_dir, 
                                     show_plot=False, plot_poly=False, yaxis_label=y_axis_label)

        # insert all polynomial (n=3) terms into cubic spline df for information purposes
        splines_df.insert(0, 'PolyReg x^3', polyreg_df['x^2'].values)
        splines_df.insert(1, 'PolyReg x^2', polyreg_df['x^3'].values)
        splines_df.insert(2, 'PolyReg x', polyreg_df['x'].values)
        splines_df.insert(3, 'PolyReg Intercept', polyreg_df['Intercept'].values)


    # obtain ARIMA forecast results
    arima_df = arima_forecasts(daily_df, save=True, save_dir=arima_dir, 
                               show_plot=False, yaxis_label=y_axis_label)

    return polyreg_df, splines_df, arima_df


def form_final_dataframe(daily_df, start_date, prediction_df, area_code_mapping=None):
    """ Form final dataframe from the original daily cases dataframe, the start date,
        and chosen prediction values from the passed model dataframe (cubic spline best) 

    Parameters
    ----------
    daily_df : Pandas DataFrame
        Pandas dataframe with the time series of cases for each location
    start_date : string
        String representing the first date within daily_df, for example: '2020-03-10'
    prediction_df : python dict
        DataFrame containing the model values to be used for the forecasted days ahead of 
        the last date in daily_df.

    Returns
    -------
    Pandas DataFrame
        Final Pandas DataFrame with original daily cases, predictions, and prediction gradients.
    """
    temp_df = daily_df.T.copy()
    date_index = pd.date_range(start=start_date, 
                           periods=daily_df.shape[0]).strftime('%d-%b-%y')
    temp_df.columns = date_index
    temp_df.index.rename('Location', inplace=True)

    preds = prediction_df[['+1 Day Forecast', '+2 Day Forecast', 
                           '+3 Day Forecast', '+4 Day Forecast', 
                           '+5 Day Forecast', 'Forecast Gradient']]

    # form list of column names for final df
    column_names = list(temp_df.columns)
    column_names = column_names.extend(preds.columns)

    # set same index and join dataframes together to make final
    preds.index = temp_df.index
    final_df = temp_df.join(preds)

    # if area codes passed, add column and sort
    if area_code_mapping:
        return map_insert_and_sort_df(final_df, 'Area code', area_code_mapping)

    return final_df



def map_insert_and_sort_df(data_df, new_col_name, mapping_values):
    """ Insert a new column of mapped values into given dataframe, followed by making this
        the index and sorting the dataframe accordingly. The original index will by the
        second column in the resulting dataframe 

    Parameters
    ----------
    data_df : Pandas DataFrame
        Pandas dataframe with the keys in mapping_values as the index
    new_col_name : string
        String containing the name of the new dataframe column to form and set as the index
    mapping_values : python dict
        Python dictionary containing keys (matching data_df index values) and mapped values
        for each unique key.

    Returns
    -------
    Pandas DataFrame
        Pandas DataFrame sorted by new column index, with the old index as second column.
    """
    data_df[new_col_name] = data_df.index.map(mapping_values)
    data_df.reset_index(inplace=True)
    data_df.set_index(new_col_name, inplace=True)
    data_df.sort_index(inplace=True)
    return data_df


def save_xls(df_list, df_names, xls_path, plot_dirs, plot_subdirs, cell_mapping, import_plots=False):
    """ Save a list of dataframes to a single Excel file """
    with ExcelWriter(xls_path) as writer:
        for n, df in enumerate(df_list):
            df.to_excel(writer, f"{df_names[n]}")
        
        # if selected, import all plots into the spreadsheet
        if import_plots:
            
            workbook  = writer.book
            
            # iterate through all plot dirs and insert into Excel
            for location, loc_dir in plot_dir_tuples:
                for model_name, model_dir in plot_subdir_tuples:
                    image_list = [img for img in sorted(os.listdir(os.path.join(loc_dir, model_dir))) 
                                  if img.endswith('.png')]
                
                    # create new sheet with given name
                    worksheet = workbook.add_worksheet(f"{location} {model_name} Plots")
                    
                    # insert all images into new sheet
                    for ind, image in enumerate(image_list):
                        img_path = os.path.join(loc_dir, model_dir, image)
                        worksheet.insert_image(cell_mapping[ind], img_path)


def format_cells(sheet, start_col, end_col, num_rows, 
                 head_fill=None, cell_fill=None, width=15, cell_style=None):
    """ Format excel table cells using Openpyxl """
    
    for x in range(start_col, end_col):
        
        # get column letter
        col_letter = get_column_letter(x)
    
        # set width of current column
        sheet.column_dimensions[col_letter].width=width
    
        if head_fill:
            # format header fill color
            sheet[f'{col_letter}1'].fill = head_fill
    
        for row in range(2, num_rows + 1):
            # format preds cells fill color
            if cell_style:
                sheet[f'{col_letter}{row}'].style = cell_style
            if cell_fill:
                sheet[f'{col_letter}{row}'].fill = cell_fill


# sequence to run if script manually run using python command line
if __name__ == "__main__":

    # create location dirs if not already existing
    for location in LOCATIONS:
        create_dirs(location)

    # form dataframes for all location datasets
    wales_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'wales-regions-lab-confirmed-per-pop.csv'), 
                             index_col='Specimen date')
    scotland_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'scotland-regions-confirmed-per-pop.csv'),
                                parse_dates=['Date'], index_col='Date')
    england_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'england-areas-lab-confirmed-per-pop.csv'),
                               index_col='Specimen date')
    ireland_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'ireland-regions-confirmed-per-pop.csv'),
                               index_col='Date')
    ni_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'ni-regions-confirmed-per-pop.csv'),
                               index_col='Date')
    other_daily_df = pd.read_csv(os.path.join(DATA_SOURCE_DIR, 'other-regions-confirmed-per-pop.csv'),
                               index_col='Date')

    
    # Scotland has a huge data dump on 15 Jun 20 - temporarily correct this for modelling
    scotland_dump_date_data = scotland_daily_df.loc['2020-06-15'].copy()

    # set the data to NaN, and interpolate suitable values using linear interpolation
    scotland_daily_df.loc['2020-06-15'] = np.nan
    scotland_daily_df.interpolate(method='slinear', inplace=True)

    # obtain Scotland areas for use later
    SCOTLAND_AREAS = list(scotland_daily_df.columns)

    # NI also has a large data dump / correction on 25-Jun-20
    ni_dump_date_data = ni_daily_df.loc['2020-06-25'].copy()
    ni_daily_df.loc['2020-06-25'] = np.nan
    ni_daily_df.interpolate(method='slinear', inplace=True)
    NI_AREAS = list(ni_daily_df.columns)


    # create a dictionary containing all location polynomial results
    polyreg_preds_dict = dict()

    # get all sets of results for our locations
    wales_polyreg_df, wales_splines_df, wales_arima_df = get_model_results(wales_daily_df, WALES_DIR, Y_AXIS_LABEL)
    scotland_polyreg_df, scotland_splines_df, scotland_arima_df = get_model_results(scotland_daily_df, SCOTLAND_DIR, Y_AXIS_LABEL)
    england_polyreg_df, england_splines_df, england_arima_df = get_model_results(england_daily_df, ENGLAND_DIR, Y_AXIS_LABEL)
    ireland_polyreg_df, ireland_splines_df, ireland_arima_df = get_model_results(ireland_daily_df, IRELAND_DIR, Y_AXIS_LABEL)
    ni_polyreg_df, ni_splines_df, ni_arima_df = get_model_results(ni_daily_df, NI_DIR, Y_AXIS_LABEL)
    other_polyreg_df, other_splines_df, other_arima_df = get_model_results(other_daily_df, OTHER_DIR, Y_AXIS_LABEL)

    # import area name to area code mappings to append this to our final data
    area_mappings = pd.read_csv(os.path.join(DATA_SOURCE_DIR, AREA_MAPPINGS_FILENAME), 
                                     index_col='name').to_dict()['code']

    # form final dataframe for each location
    final_wales_df = form_final_dataframe(wales_daily_df, WALES_START_DATE, wales_splines_df, area_mappings)
    final_scotland_df = form_final_dataframe(scotland_daily_df, SCOTLAND_START_DATE, scotland_splines_df, area_mappings)
    final_england_df = form_final_dataframe(england_daily_df, ENGLAND_START_DATE, england_splines_df, area_mappings)
    final_ireland_df = form_final_dataframe(ireland_daily_df, IRELAND_START_DATE, ireland_splines_df, area_mappings)
    final_ni_df = form_final_dataframe(ni_daily_df, NI_START_DATE, ni_splines_df, area_mappings)
    final_other_df = form_final_dataframe(other_daily_df, OTHER_START_DATE, other_splines_df, area_mappings)

    # form final df list to insert into Excel file
    final_df_list = [final_england_df, final_scotland_df, final_wales_df, 
                     final_ireland_df, final_ni_df, final_other_df,
                     england_splines_df, scotland_splines_df, wales_splines_df, 
                     ireland_splines_df,  ni_splines_df, other_splines_df,
                     england_arima_df, scotland_arima_df, wales_arima_df, 
                     ireland_arima_df,  ni_arima_df, other_arima_df]

    # form list of final dataframe names to use for each excel sheet name
    final_df_names = ['England Cases Per Pop', 'Scotland Cases Per Pop', 'Wales Cases Per Pop', 
                      'Ireland Cases Per Pop', 'NI Cases Per Pop', 'Other Cases Per Pop',
                      'England CubicSpline Model', 'Scotland CubicSpline Model', 'Wales CubicSpline Model',
                      'Ireland CubicSpline Model', 'NI CubicSpline Model', 'Other CubicSpline Model',
                      'England ARIMA Model', 'Scotland ARIMA Model', 'Wales ARIMA Model',
                      'Ireland ARIMA Model', 'NI ARIMA Model', 'Other ARIMA Model']

    # set tuples of names and dir locations to allow fetching of all saved images from each dir
    plot_dir_tuples = [('England', ENGLAND_DIR), ('Scotland', SCOTLAND_DIR), ('Wales', WALES_DIR), 
                       ('Ireland', IRELAND_DIR), ('NI', NI_DIR), ('Other', OTHER_DIR)]

    plot_subdir_tuples = [('CubicSpline', CUBIC_SPLINE_FOLDER), ('ARIMA', ARIMA_FOLDER)]

    # create an appropriate cell mapping for plots into excel worksheet
    # this works well for plot size (7,5)
    number_range = [x for x in range(0, 200)]
    cell_strings = [f'A{x} K{x} U{x}' for x in range(1, 1500, 24)]
    cell_list = " ".join(cell_strings).split()
    cell_mapping = {a:b for a,b in zip(number_range, cell_list)}

    # save all of our results above into a single DataFrame
    save_xls(final_df_list, final_df_names, os.path.join(os.getcwd(), 'temp_excelfile.xlsx'), 
                plot_dir_tuples, plot_subdir_tuples, cell_mapping, import_plots=True)

    # set up styles and formatting for openpyxl to neaten up our excel file
    first_cols_head = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    first_cols_fill = PatternFill(fill_type='solid', start_color='F2F2F2', end_color='F2F2F2')

    time_head_fill = PatternFill(fill_type='solid', start_color='4C7DB7', end_color='4C7DB7')

    forecast_header_fill = PatternFill(fill_type='solid', start_color='E26B0A', end_color='E26B0A')
    forecasts_fill = PatternFill(fill_type='solid', start_color='FDE9D9', end_color='FDE9D9')

    gradient_header_fill = PatternFill(fill_type='solid', start_color='7030A0', end_color='7030A0')
    gradient_fill = PatternFill(fill_type='solid', start_color='FDE9D9', end_color='FDE9D9')

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # format header cells
    header = NamedStyle(name="header")
    header.font = Font(bold=True, color='FFFFFFFF')
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.height = 25
    header.fill = time_head_fill
    header.border = thin_border

    # format numeric cells
    number_cell = NamedStyle(name="number_cell")
    number_cell.alignment = Alignment(horizontal="center", vertical="center")
    number_cell.number_format = '#,##0.000'
    number_cell.border = thin_border

    # format generic text cells
    text_cell = NamedStyle(name="text_cell")
    text_cell.alignment = Alignment(horizontal="center", vertical="center")
    text_cell.border = thin_border

    # open excel file and ammend styling as required
    workbook = openpyxl.load_workbook('temp_excelfile.xlsx')

    # sequentially format all sheets containing tables
    for sheet_name in final_df_names:
        current_sheet = workbook[sheet_name]

        print(f"Currently editing sheet {sheet_name}")
    
        # freeze the first column of the sheet
        current_sheet.freeze_panes = 'C2'
    
        # format all header cells
        header_row = current_sheet[1]
        for cell in header_row:
            cell.style = header
    
        # set height of header row
        current_sheet.row_dimensions[1].height = 25
    
        num_rows = current_sheet.max_row
        num_cols = current_sheet.max_column

        # format first two columns and cells
        format_cells(current_sheet, 1, 3, num_rows, head_fill=first_cols_head, 
                    cell_fill=first_cols_fill, width=FIRST_COL_WIDTHS, cell_style=text_cell)

        # format forecast columns and cells
        format_cells(current_sheet, num_cols-5, num_cols, num_rows, head_fill=forecast_header_fill, 
                    cell_fill=forecasts_fill, width=FORECAST_WIDTHS, cell_style=number_cell)

        # format gradient columns and cells
        format_cells(current_sheet, num_cols, num_cols + 1, num_rows, head_fill=gradient_header_fill,
                    cell_fill=None, width=FORECAST_WIDTHS, cell_style=number_cell)

        # format bulk of numeric cells
        format_cells(current_sheet, 3, num_cols - 5, num_rows, head_fill=None, cell_fill=None,
                    width=TIME_CELL_WIDTHS, cell_style=number_cell)

        # modify width of first col
        current_sheet.column_dimensions['A'].width = 20
    
    # sort sheets into final desired order before saving
    workbook._sheets.sort(key=lambda sheet: FINAL_SHEET_ORDER[sheet.title])

    # save changes to workbook
    workbook.save(FILE_SAVE_NAME)

    # delete all temporary files and data
    if os.path.isfile("temp_excelfile.xlsx"):
        os.remove("temp_excelfile.xlsx")